import logging
import time

from openpyxl import Workbook, load_workbook

import config

logger = logging.getLogger("work_order_automation")


class ExcelWriteError(Exception):
    """Raised when the Excel file cannot be written to."""


def _find_last_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=1).value not in (None, ""):
            return row
    return 0


def append_work_order(row_data):
    """Append one row to work_orders.xlsx.

    Args:
        row_data: dict with keys work_order, item_code, description, qty.

    Raises:
        ExcelWriteError: if the file stays locked (open in Excel) or cannot be written.
    """
    if not config.EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(config.EXCEL_HEADERS)
        wb.save(config.EXCEL_FILE)
        logger.info("Created %s with header row", config.EXCEL_FILE)

    for attempt in range(1, config.EXCEL_MAX_RETRIES + 1):
        try:
            wb = load_workbook(config.EXCEL_FILE)
            ws = wb.active
            last_row = _find_last_row(ws)
            next_sr_no = 1 if last_row < 1 else ws.cell(row=last_row, column=1).value
            if next_sr_no is None:
                next_sr_no = last_row
            try:
                next_sr_no = int(next_sr_no) + 1
            except (TypeError, ValueError):
                next_sr_no = last_row + 1

            ws.append(
                [
                    next_sr_no,
                    row_data["work_order"],
                    row_data["item_code"],
                    row_data["description"],
                    row_data["qty"],
                    "",  # Cat - always blank
                    "",  # Status - always blank
                ]
            )
            wb.save(config.EXCEL_FILE)
            logger.info(
                "Excel row appended: Sr No %s, Work Order %s",
                next_sr_no,
                row_data["work_order"],
            )
            return next_sr_no
        except PermissionError:
            logger.warning(
                "Cannot save Excel file (attempt %s/%s) - is it open in Excel? "
                "Please close %s.",
                attempt,
                config.EXCEL_MAX_RETRIES,
                config.EXCEL_FILE,
            )
            if attempt < config.EXCEL_MAX_RETRIES:
                time.sleep(config.EXCEL_RETRY_DELAY_SECONDS)
        except Exception as exc:
            raise ExcelWriteError(f"Could not write to Excel file: {exc}")

    raise ExcelWriteError(
        f"Excel file is still locked after {config.EXCEL_MAX_RETRIES} attempts. "
        f"PLEASE CLOSE the file {config.EXCEL_FILE} in Excel, then drop the photo again."
    )
